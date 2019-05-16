from openpyxl import load_workbook


# 加载 excel 文件：默认可读写
wb = load_workbook("template.xlsm")
# 获得所有数据表的名称
sheets = wb.get_sheet_names()
# 获取凭证库表
ivoucher = wb.get_sheet_by_name("凭证库")
# 获取行与列
rows = ivoucher.max_row
cols = ivoucher.max_column


"""
code: 科目代码
digest: 摘要
ledger: 总账科目
inventory：明细科目
debit：借金额
credit：贷金额
voucher：凭证号
attachments：附件数
"""
col_ids = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
voucher = int(input("请输入凭证号："))
data = {
    'code': [],
    'digest': [],
    'ledger': [],
    'inventory': [],
    'debit': [],
    'credit': [],
    'voucher': voucher,
    'attachments': [],
    }

# 获取编制凭证所需数据
for row in ivoucher.rows:
    row_list = list(row)
    if row_list[8].value == voucher:
        data['code'].append(row_list[0].value)
        data['digest'].append(row_list[1].value)
        data['ledger'].append(row_list[2].value)
        if row_list[3].value is None:
            data['inventory'].append('')
        else:
            data['inventory'].append(row_list[3].value)
        data['debit'].append(row_list[6].value)
        data['credit'].append(row_list[7].value)
        if row_list[9].value is None:
            data['attachments'].append(0)
        else:
            data['attachments'].append(row_list[9].value)

data_list = [['科目代码', '摘要', '总账科目',
             '明细科目', '借金额', '贷金额', '附件数']]

for i in range(len(data['code'])):
    line = [[data['code'][i], data['digest'][i], data['ledger'][i],
            data['inventory'][i], data['debit'][i], data['credit'][i],
            data['attachments'][i]]]
    data_list.extend(line)

# 生成新的凭证表
new_voucher = wb.create_sheet("记账凭证")
for j in range(len(data_list)):
    for k in range(len(data_list[j])):
        new_voucher.cell(row=(j+1), column=(k+1)).value = data_list[k][j]

wb.save(r'template.xlsx')
